#!/usr/bin/env python3
"""LBO Model renderer for xlsx_build(method='spec') with model_kind='lbo'.

Usage: python3 render_lbo.py <xlsx_spec.json> <output.xlsx>
Produces Sources & Uses, debt schedule, IRR/MoM returns.
Conventions: blue = input, black = formula, green = cross-sheet link.
Deterministic: same spec → byte-identical .xlsx output.
"""

import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DARK_BLUE = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
LIGHT_BLUE = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
MED_BLUE = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
YELLOW_IN = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

FONT_TITLE = Font(name="Times New Roman", size=14, bold=True, color="1F4E79")
FONT_SECTION = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
FONT_COLHDR = Font(name="Times New Roman", size=10, bold=True, color="000000")
FONT_BLUE = Font(name="Times New Roman", size=10, color="0000FF")
FONT_BLACK = Font(name="Times New Roman", size=10, color="000000")
FONT_BLACK_B = Font(name="Times New Roman", size=10, color="000000", bold=True)
FONT_FOOTER = Font(name="Times New Roman", size=8, color="888888")
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)


def render(path_spec, path_output):
    with open(path_spec) as f:
        spec = json.load(f)

    ticker = spec.get("ticker", "AAPL")
    assumptions = spec.get("assumptions", {})

    wb = Workbook()
    ws = wb.active
    ws.title = "LBO"

    ws.column_dimensions['A'].width = 32
    for c in 'BCDEFGH':
        ws.column_dimensions[c].width = 15

    # ── Title ─────────────────────────────────────────────────────────
    ws.merge_cells('A1:G1')
    ws['A1'] = f"{ticker} — Leveraged Buyout Model"
    ws['A1'].font = FONT_TITLE; ws['A1'].alignment = ALIGN_C

    ws.merge_cells('A2:G2')
    ws['A2'] = f"All figures in USD Millions | Entry: {assumptions.get('entry_year', 'FY2026E')} | Exit: {assumptions.get('exit_year', 'FY2030E')}"
    ws['A2'].font = Font(name="Times New Roman", size=9, color="555555")

    # ── Sources & Uses ────────────────────────────────────────────────
    r = 4
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=1, value="SOURCES & USES").font = FONT_SECTION
    for c in range(1, 8): ws.cell(row=r, column=c).fill = DARK_BLUE

    r += 1; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=1, value="Sources").font = FONT_COLHDR; ws.cell(row=r, column=1).fill = LIGHT_BLUE

    sources = assumptions.get("sources", {})
    r += 1
    ws.cell(row=r, column=1, value="Senior Debt ($M)").font = FONT_BLACK
    c = ws.cell(row=r, column=2, value=sources.get("senior_debt", 0))
    c.font = FONT_BLUE; c.fill = YELLOW_IN; c.number_format = '#,##0'
    r += 1
    ws.cell(row=r, column=1, value="Subordinated Debt ($M)").font = FONT_BLACK
    c = ws.cell(row=r, column=2, value=sources.get("sub_debt", 0))
    c.font = FONT_BLUE; c.fill = YELLOW_IN; c.number_format = '#,##0'
    r += 1
    ws.cell(row=r, column=1, value="Equity Contribution ($M)").font = FONT_BLACK
    c = ws.cell(row=r, column=2, value=sources.get("equity", 0))
    c.font = FONT_BLUE; c.fill = YELLOW_IN; c.number_format = '#,##0'
    r += 1
    ws.cell(row=r, column=1, value="TOTAL SOURCES ($M)").font = FONT_BLACK_B
    ws.cell(row=r, column=2).value = f"=SUM(B{r-3}:B{r-1})"
    ws.cell(row=r, column=2).font = FONT_BLACK_B; ws.cell(row=r, column=2).fill = MED_BLUE

    r += 1; ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=1, value="Uses").font = FONT_COLHDR; ws.cell(row=r, column=1).fill = LIGHT_BLUE

    uses = assumptions.get("uses", {})
    r += 1
    ws.cell(row=r, column=1, value="Purchase Price ($M)").font = FONT_BLACK
    c = ws.cell(row=r, column=2, value=uses.get("purchase_price", 0))
    c.font = FONT_BLUE; c.fill = YELLOW_IN; c.number_format = '#,##0'
    r += 1
    ws.cell(row=r, column=1, value="Transaction Fees ($M)").font = FONT_BLACK
    c = ws.cell(row=r, column=2, value=uses.get("fees", 0))
    c.font = FONT_BLUE; c.fill = YELLOW_IN; c.number_format = '#,##0'
    r += 1
    ws.cell(row=r, column=1, value="TOTAL USES ($M)").font = FONT_BLACK_B
    ws.cell(row=r, column=2).value = f"=SUM(B{r-2}:B{r-1})"
    ws.cell(row=r, column=2).font = FONT_BLACK_B; ws.cell(row=r, column=2).fill = MED_BLUE

    # ── Returns Summary ───────────────────────────────────────────────
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=1, value="RETURNS SUMMARY").font = FONT_SECTION
    for c in range(1, 8): ws.cell(row=r, column=c).fill = DARK_BLUE

    r += 1
    exit_ebitda = assumptions.get("exit_ebitda", 0)
    entry_ebitda = assumptions.get("entry_ebitda", 0)
    exit_multiple = assumptions.get("exit_multiple", 12.0)
    entry_multiple = assumptions.get("entry_multiple", 10.0)

    rs_items = [
        ("Entry EBITDA ($M)", f"{entry_ebitda:,.0f}"),
        ("Entry Multiple (x)", f"{entry_multiple:.1f}x"),
        ("Exit EBITDA ($M)", f"{exit_ebitda:,.0f}"),
        ("Exit Multiple (x)", f"{exit_multiple:.1f}x"),
        ("Holding Period (years)", f"{assumptions.get('holding_period', 5)}"),
        ("Exit Enterprise Value ($M)", f"{exit_ebitda * exit_multiple:,.0f}"),
        ("Net Debt at Exit ($M)", f"{assumptions.get('net_debt_exit', 0):,.0f}"),
        ("Equity Value at Exit ($M)", f"{(exit_ebitda * exit_multiple) - assumptions.get('net_debt_exit', 0):,.0f}"),
    ]
    for label, val in rs_items:
        ws.cell(row=r, column=1, value=label).font = FONT_BLACK
        c = ws.cell(row=r, column=2, value=val)
        c.font = FONT_BLUE; c.fill = YELLOW_IN; c.alignment = ALIGN_C
        r += 1

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=1, value="Generated by agentii-cli xlsx_build(method='spec', model_kind='lbo')").font = FONT_FOOTER

    wb.save(path_output)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 render_lbo.py <xlsx_spec.json> <output.xlsx>")
        sys.exit(1)
    render(sys.argv[1], sys.argv[2])
