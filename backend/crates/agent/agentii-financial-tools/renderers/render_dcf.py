#!/usr/bin/env python3
"""DCF model renderer for xlsx_build(method='spec') with model_kind='dcf'.

Usage: python3 render_dcf.py <xlsx_spec.json> <output.xlsx>
Input: xlsx_spec JSON file path (spec 023 xlsx_spec.schema.json conformant)
Output: .xlsx file at output path

Convention: blue font = hardcoded input, black font = formula, green = cross-sheet link.
Deterministic: same spec → byte-identical .xlsx output.
"""

import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Style constants (Anthropic FSI conventions) ────────────────────────
DARK_BLUE = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
LIGHT_BLUE = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
MED_BLUE = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
LIGHT_GREY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
YELLOW_IN = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

FONT_TITLE = Font(name="Times New Roman", size=14, bold=True, color="1F4E79")
FONT_SECTION = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
FONT_COLHDR = Font(name="Times New Roman", size=10, bold=True, color="000000")
FONT_BLUE = Font(name="Times New Roman", size=10, color="0000FF")
FONT_BLACK = Font(name="Times New Roman", size=10, color="000000")
FONT_BLACK_B = Font(name="Times New Roman", size=10, color="000000", bold=True)
FONT_GREEN = Font(name="Times New Roman", size=10, color="008000")
FONT_FOOTER = Font(name="Times New Roman", size=8, color="888888")

ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
THICK = Border(left=Side('medium'), right=Side('medium'), top=Side('medium'), bottom=Side('medium'))


def render(path_spec, path_output):
    with open(path_spec) as f:
        spec = json.load(f)

    model = spec.get("model_kind", "dcf")
    ticker = spec.get("ticker", "AAPL")
    title = spec.get("title", f"{ticker} DCF Model")

    wb = Workbook()
    ws = wb.active
    ws.title = "DCF"

    # Column widths
    ws.column_dimensions['A'].width = 32
    for c in 'BCDEFGH':
        ws.column_dimensions[c].width = 15

    # ── Title ─────────────────────────────────────────────────────────
    ws.merge_cells('A1:H1')
    ws['A1'] = f"{title} — Discounted Cash Flow Valuation"
    ws['A1'].font = FONT_TITLE
    ws['A1'].alignment = ALIGN_C

    ws.merge_cells('A2:H2')
    ws['A2'] = f"Model: {model} | Ticker: {ticker} | All figures in USD Millions unless noted"
    ws['A2'].font = Font(name="Times New Roman", size=9, color="555555")
    ws['A2'].alignment = ALIGN_C

    # ── Assumptions Section ───────────────────────────────────────────
    r = 4
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=1, value="KEY ASSUMPTIONS").font = FONT_SECTION
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = DARK_BLUE

    assumptions = spec.get("assumptions", {})
    r += 1
    for label, key in [
        ("Revenue Growth (FY1-FY5)", "revenue_growth"),
        ("EBITDA Margin (Terminal)", "ebitda_margin"),
        ("WACC (%)", "wacc"),
        ("Terminal Growth (%)", "terminal_growth"),
        ("Terminal EV/EBITDA (x)", "terminal_ev_ebitda"),
        ("Tax Rate (%)", "tax_rate"),
    ]:
        val = assumptions.get(key, 0)
        if isinstance(val, list):
            val_str = ", ".join(f"{v*100:.0f}%" for v in val[:5])
        elif isinstance(val, float) and val < 1:
            val_str = f"{val*100:.1f}%"
        else:
            val_str = str(val)
        ws.cell(row=r, column=1, value=label).font = FONT_BLACK
        c = ws.cell(row=r, column=2, value=val_str)
        c.font = FONT_BLUE
        c.fill = YELLOW_IN
        c.alignment = ALIGN_C
        r += 1

    # ── Projection Years ──────────────────────────────────────────────
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=1, value="PROJECTED CASH FLOWS").font = FONT_SECTION
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = DARK_BLUE

    r += 1
    years = spec.get("projection_years", ["FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"])
    ws.cell(row=r, column=1, value="").font = FONT_COLHDR
    for i, yr in enumerate(years[:5]):
        c = ws.cell(row=r, column=3+i, value=yr)
        c.font = FONT_COLHDR
        c.fill = LIGHT_BLUE
        c.alignment = ALIGN_C
    colhdr_row = r

    # ── Revenue ───────────────────────────────────────────────────────
    r += 1
    revenue = spec.get("revenue", [])
    ws.cell(row=r, column=1, value="Revenue ($M)").font = FONT_BLACK_B
    for i, rev in enumerate(revenue[:5]):
        c = ws.cell(row=r, column=3+i, value=rev)
        c.font = FONT_BLUE if i == 0 else FONT_BLACK
        c.number_format = '#,##0'
        c.alignment = ALIGN_C

    # ── EBITDA ────────────────────────────────────────────────────────
    r += 1
    ebitda_margins = assumptions.get("ebitda_margin", 0.35)
    if not isinstance(ebitda_margins, list):
        ebitda_margins = [ebitda_margins] * 5
    ws.cell(row=r, column=1, value="EBITDA ($M)").font = FONT_BLACK_B
    for i in range(min(5, len(revenue))):
        c = ws.cell(row=r, column=3+i)
        c.value = f"=E{r-1}*{ebitda_margins[i]}" if i == 0 else f"={get_column_letter(3+i)}{r-1}*{ebitda_margins[i]}"
        c.font = FONT_BLACK
        c.number_format = '#,##0'
        c.alignment = ALIGN_C

    # ── D&A ───────────────────────────────────────────────────────────
    r += 1
    da_pct = assumptions.get("da_pct", 0.025)
    ws.cell(row=r, column=1, value="(-) D&A ($M)").font = FONT_BLACK
    for i in range(min(5, len(revenue))):
        c = ws.cell(row=r, column=3+i)
        c.value = f"=E{r-2}*{da_pct}" if i == 0 else f"={get_column_letter(3+i)}{r-2}*{da_pct}"
        c.font = FONT_BLACK
        c.number_format = '#,##0'
        c.alignment = ALIGN_C

    # ── EBIT ──────────────────────────────────────────────────────────
    r += 1
    ws.cell(row=r, column=1, value="EBIT ($M)").font = FONT_BLACK_B
    for i in range(min(5, len(revenue))):
        c = ws.cell(row=r, column=3+i)
        c.value = f"=E{r-1}-E{r}" if i == 0 else f"={get_column_letter(3+i)}{r-1}-{get_column_letter(3+i)}{r}"
        c.font = FONT_BLACK_B
        c.fill = MED_BLUE
        c.number_format = '#,##0'
        c.alignment = ALIGN_C

    ebit_row = r

    # ── Summary ───────────────────────────────────────────────────────
    r = ebit_row + 3
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=1, value="VALUATION SUMMARY").font = FONT_SECTION
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = DARK_BLUE

    r += 1
    wacc = assumptions.get("wacc", 0.105)
    term_g = assumptions.get("terminal_growth", 0.03)
    ev_ebitda = assumptions.get("terminal_ev_ebitda", 22)

    ws.cell(row=r, column=1, value="WACC").font = FONT_BLACK
    ws.cell(row=r, column=2, value=f"{wacc*100:.1f}%").font = FONT_BLUE
    r += 1
    ws.cell(row=r, column=1, value="Terminal Growth Rate").font = FONT_BLACK
    ws.cell(row=r, column=2, value=f"{term_g*100:.1f}%").font = FONT_BLUE
    r += 1
    ws.cell(row=r, column=1, value="Terminal EV/EBITDA").font = FONT_BLACK
    ws.cell(row=r, column=2, value=f"{ev_ebitda:.1f}x").font = FONT_BLUE

    # ── Footer ────────────────────────────────────────────────────────
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=1, value=f"Generated by agentii-cli xlsx_build(method='spec', model_kind='{model}')").font = FONT_FOOTER
    ws.cell(row=r, column=1).alignment = ALIGN_C

    wb.save(path_output)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 render_dcf.py <xlsx_spec.json> <output.xlsx>")
        sys.exit(1)
    render(sys.argv[1], sys.argv[2])
