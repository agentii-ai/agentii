#!/usr/bin/env python3
"""Comparable Company Analysis renderer for xlsx_build(method='spec') with model_kind='comps'.

Usage: python3 render_comps.py <xlsx_spec.json> <output.xlsx>
Conventions: blue = input, black = formula, green = cross-sheet link.
Deterministic: same spec → byte-identical .xlsx output.
"""

import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DARK_BLUE = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
LIGHT_BLUE = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
MED_BLUE = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
LIGHT_GREY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
YELLOW_IN = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

FONT_TITLE = Font(name="Times New Roman", size=14, bold=True, color="1F4E79")
FONT_SECTION = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
FONT_COLHDR = Font(name="Times New Roman", size=10, bold=True, color="000000")
FONT_BLUE = Font(name="Times New Roman", size=10, color="0000FF")
FONT_BLACK = Font(name="Times New Roman", size=10, color="000000")
FONT_BLACK_B = Font(name="Times New Roman", size=10, color="000000", bold=True)
FONT_FOOTER = Font(name="Times New Roman", size=8, color="888888")
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
THICK = Border(left=Side('medium'), right=Side('medium'), top=Side('medium'), bottom=Side('medium'))


def render(path_spec, path_output):
    with open(path_spec) as f:
        spec = json.load(f)

    ticker = spec.get("ticker", "AAPL")
    peers = spec.get("peers", ["MSFT", "GOOGL", "AMZN", "META", "NVDA"])
    companies = [ticker] + [p for p in peers if p != ticker]

    wb = Workbook()
    ws = wb.active
    ws.title = "Comps"

    ws.column_dimensions['A'].width = 22
    for c in 'BCDEFGHIJKL':
        ws.column_dimensions[c].width = 15

    # ── Title ─────────────────────────────────────────────────────────
    ws.merge_cells('A1:L1')
    ws['A1'] = f"TECHNOLOGY — COMPARABLE COMPANY ANALYSIS"
    ws['A1'].font = FONT_TITLE
    ws['A1'].alignment = ALIGN_C
    ws.merge_cells('A2:L2')
    ws['A2'] = f"Target: {ticker} | Peers: {', '.join(peers)} | All figures in USD Billions"
    ws['A2'].font = Font(name="Times New Roman", size=9, color="555555")

    # ── Operating Metrics ─────────────────────────────────────────────
    r = 4
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=1, value="OPERATING STATISTICS & FINANCIAL METRICS").font = FONT_SECTION
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = DARK_BLUE

    r += 1
    op_headers = ["Company", "Revenue (LTM)", "Rev Growth %", "Gross Margin %", "EBITDA (LTM)", "EBITDA Margin %"]
    for i, h in enumerate(op_headers):
        ws.cell(row=r, column=1+i, value=h).font = FONT_COLHDR
        ws.cell(row=r, column=1+i).fill = LIGHT_BLUE
        ws.cell(row=r, column=1+i).alignment = ALIGN_C

    op_data = spec.get("operating_metrics", {})
    r += 1
    data_start = r
    for co in companies:
        d = op_data.get(co, {})
        ws.cell(row=r, column=1, value=co).font = FONT_BLACK_B if co == ticker else FONT_BLACK
        c2 = ws.cell(row=r, column=2, value=d.get("revenue", 0))
        c2.font = FONT_BLUE; c2.number_format = '#,##0.0'; c2.alignment = ALIGN_C
        c3 = ws.cell(row=r, column=3, value=d.get("revenue_growth", 0))
        c3.font = FONT_BLUE; c3.number_format = '0.0%'; c3.alignment = ALIGN_C
        c4 = ws.cell(row=r, column=4, value=d.get("gross_margin", 0))
        c4.font = FONT_BLUE; c4.number_format = '0.0%'; c4.alignment = ALIGN_C
        c5 = ws.cell(row=r, column=5, value=d.get("ebitda", 0))
        c5.font = FONT_BLUE; c5.number_format = '#,##0.0'; c5.alignment = ALIGN_C
        c6 = ws.cell(row=r, column=6, value=d.get("ebitda_margin", 0))
        c6.font = FONT_BLUE; c6.number_format = '0.0%'; c6.alignment = ALIGN_C
        r += 1
    data_end = r - 1

    # ── Statistics ────────────────────────────────────────────────────
    r += 1
    for label, func in [("Maximum", "MAX"), ("75th Percentile", "QUARTILE.INC"), ("Median", "MEDIAN"), ("25th Percentile", "QUARTILE.INC"), ("Minimum", "MIN")]:
        ws.cell(row=r, column=1, value=label).font = FONT_BLACK_B
        ws.cell(row=r, column=1).fill = LIGHT_GREY
        for col in [3, 4, 6]:
            cl = get_column_letter(col)
            if func == "QUARTILE.INC":
                ws.cell(row=r, column=col).value = f"=QUARTILE.INC({cl}{data_start}:{cl}{data_end},{4 if label.startswith('75') else 1})"
            else:
                ws.cell(row=r, column=col).value = f"={func}({cl}{data_start}:{cl}{data_end})"
            ws.cell(row=r, column=col).font = FONT_BLACK
            ws.cell(row=r, column=col).number_format = '0.0%'
            ws.cell(row=r, column=col).fill = LIGHT_GREY
        r += 1

    # ── Valuation Multiples ───────────────────────────────────────────
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    ws.cell(row=r, column=1, value="VALUATION MULTIPLES").font = FONT_SECTION
    for c in range(1, 13):
        ws.cell(row=r, column=c).fill = DARK_BLUE

    r += 1
    val_headers = ["Company", "Market Cap", "Enterprise Value", "EV/Revenue", "EV/EBITDA", "P/E"]
    val_cols = [1, 8, 9, 10, 11, 12]
    for i, (h, col) in enumerate(zip(val_headers, val_cols)):
        ws.cell(row=r, column=col, value=h).font = FONT_COLHDR
        ws.cell(row=r, column=col).fill = LIGHT_BLUE

    val_data = spec.get("valuation_multiples", {})
    r += 1
    val_start = r
    for co in companies:
        d = val_data.get(co, {})
        ws.cell(row=r, column=1, value=co).font = FONT_BLACK_B if co == ticker else FONT_BLACK
        for col, key in [(8, "market_cap"), (9, "enterprise_value"), (10, "ev_revenue"), (11, "ev_ebitda"), (12, "pe_ratio")]:
            c = ws.cell(row=r, column=col, value=d.get(key, 0))
            c.font = FONT_BLUE; c.alignment = ALIGN_C
            if key in ("ev_revenue", "ev_ebitda", "pe_ratio"):
                c.number_format = '0.0'
            else:
                c.number_format = '#,##0.0'
        r += 1
    val_end = r - 1

    # ── Valuation Statistics ──────────────────────────────────────────
    r += 1
    for label, func in [("Maximum", "MAX"), ("75th Percentile", "QUARTILE.INC"), ("Median", "MEDIAN"), ("25th Percentile", "QUARTILE.INC"), ("Minimum", "MIN")]:
        ws.cell(row=r, column=1, value=label).font = FONT_BLACK_B
        ws.cell(row=r, column=1).fill = LIGHT_GREY
        for col in [10, 11, 12]:
            cl = get_column_letter(col)
            if func == "QUARTILE.INC":
                ws.cell(row=r, column=col).value = f"=QUARTILE.INC({cl}{val_start}:{cl}{val_end},{4 if label.startswith('75') else 1})"
            else:
                ws.cell(row=r, column=col).value = f"={func}({cl}{val_start}:{cl}{val_end})"
            ws.cell(row=r, column=col).font = FONT_BLACK
            ws.cell(row=r, column=col).fill = LIGHT_GREY
        r += 1

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
    ws.cell(row=r, column=1, value="Generated by agentii-cli xlsx_build(method='spec', model_kind='comps')").font = FONT_FOOTER

    wb.save(path_output)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 render_comps.py <xlsx_spec.json> <output.xlsx>")
        sys.exit(1)
    render(sys.argv[1], sys.argv[2])
